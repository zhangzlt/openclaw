#!/usr/bin/env python3
"""
为非对话型智能体生成测试文件（PDF / Excel）

覆盖场景:
- 合同比对 (销售合同 + 采购合同)
- URS 需求规格书解析
- 担保/授信合同解析
- 报价单审核
- 企业信息表格自动填写 (Excel)
- PDF 脱敏打码 (含敏感信息)
"""

import os
from pathlib import Path
from fpdf import FPDF

OUT_DIR = Path(__file__).parent / "test_files"
OUT_DIR.mkdir(exist_ok=True)


class TestPDF(FPDF):
    """带标准页眉的中文 PDF"""
    def __init__(self, title: str):
        super().__init__()
        self.title_text = title
        self.add_font("CJK", "", "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc")
        self.add_font("CJK", "B", "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc")
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_font("CJK", "B", 12)
        self.cell(0, 10, self.title_text, new_x="LMARGIN", new_y="NEXT", align="C")
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

    def section(self, heading: str):
        self.set_font("CJK", "B", 11)
        self.cell(0, 8, heading, new_x="LMARGIN", new_y="NEXT")
        self.set_font("CJK", "", 10)

    def body(self, text: str):
        self.multi_cell(0, 6, text)
        self.ln(3)

    def table_row(self, key: str, value: str):
        self.set_font("CJK", "B", 10)
        self.cell(50, 7, key + "：")
        self.set_font("CJK", "", 10)
        self.cell(0, 7, value, new_x="LMARGIN", new_y="NEXT")


def generate_sales_contract():
    """生成测试销售合同 PDF [126] 智能采销合同比对 用"""
    pdf = TestPDF("销售合同")
    pdf.add_page()
    pdf.section("一、合同基本信息")
    pdf.table_row("合同编号", "XS-2026-0701-001")
    pdf.table_row("签订日期", "2026年7月1日")
    pdf.table_row("卖方（供方）", "神州数码（中国）有限公司")
    pdf.table_row("买方（需方）", "北京元力科技有限公司")
    pdf.section("二、产品与服务")
    pdf.body("1. 卖方同意向买方提供以下产品及服务：\n   - ThinkPad X1 Carbon Gen12 笔记本 × 50 台\n   - 配套三年上门保修服务\n   - 设备部署实施服务")
    pdf.section("三、价格与付款")
    pdf.table_row("合同总金额", "人民币 肆拾玖万玖仟伍佰元整 (¥499,500.00)")
    pdf.table_row("单价", "¥9,990.00/台")
    pdf.table_row("付款方式", "签约后 30 日内支付 50%，验收后 30 日内支付 50%")
    pdf.section("四、交付与验收")
    pdf.body("1. 交付时间：合同生效后 60 个工作日内。\n2. 交付地点：买方指定地点（北京市海淀区）。\n3. 验收标准：按产品规格书执行，买方应在收到货物后 10 个工作日内完成验收。")
    pdf.section("五、违约责任")
    pdf.body("1. 若卖方逾期交付，每逾期一日按未交付部分金额的 0.1% 支付违约金。\n2. 若买方逾期付款，每逾期一日按未付款金额的 0.05% 支付滞纳金。")
    pdf.section("六、签章")
    pdf.body("卖方代表（签章）：________________  买方代表（签章）：________________\n日期：2026年7月1日          日期：2026年7月1日")
    path = OUT_DIR / "sales_contract_test.pdf"
    pdf.output(str(path))
    print(f"  ✅ {path.name}")


def generate_purchase_contract():
    """生成测试采购合同 PDF [126] 用"""
    pdf = TestPDF("采购合同")
    pdf.add_page()
    pdf.section("一、合同基本信息")
    pdf.table_row("合同编号", "CG-2026-0701-088")
    pdf.table_row("签订日期", "2026年7月5日")
    pdf.table_row("买方", "神州数码（中国）有限公司")
    pdf.table_row("卖方（供应商）", "联想（北京）信息技术有限公司")
    pdf.section("二、采购产品")
    pdf.body("1. ThinkPad X1 Carbon Gen12 笔记本 × 50 台\n   规格：i7-1365U / 16GB / 512GB SSD / 14\" 2.8K OLED")
    pdf.section("三、价格条款")
    pdf.table_row("采购单价", "¥8,500.00/台")
    pdf.table_row("采购总金额", "人民币 肆拾贰万伍仟元整 (¥425,000.00)")
    pdf.table_row("付款条件", "货到验收合格后 60 日内支付全款")
    pdf.table_row("发票类型", "增值税专用发票（税率 13%）")
    pdf.section("四、交付条款")
    pdf.body("1. 交付期限：订单确认后 45 个工作日内。\n2. 运输方式：卖方负责运输，费用由卖方承担。\n3. 交货地点：买方指定仓库。")
    pdf.section("五、质量保证")
    pdf.body("1. 产品质保期为验收合格之日起 36 个月。\n2. 质保期内出现质量问题，卖方负责免费维修或更换。")
    pdf.section("六、签章")
    pdf.body("买方代表（签章）：________________  卖方代表（签章）：________________")
    path = OUT_DIR / "purchase_contract_test.pdf"
    pdf.output(str(path))
    print(f"  ✅ {path.name}")


def generate_urs_document():
    """生成测试 URS 需求规格书 [123] 售前URS解析助手 用"""
    pdf = TestPDF("用户需求规格书 (URS)")
    pdf.add_page()
    pdf.section("一、项目概述")
    pdf.body("本项目旨在为某大型零售企业构建一套智能化供应链管理系统，实现采购、库存、物流的全链路数字化管理，提升运营效率30%以上。")
    pdf.section("二、功能需求")
    pdf.body("FR-001: 系统应支持多供应商自动比价，基于历史采购数据推荐最优供应商。\nFR-002: 需实现库存实时可视化，支持最低库存预警与自动补货建议。\nFR-003: 物流跟踪模块需对接主流快递公司API（顺丰、京东、圆通）。\nFR-004: 系统需支持移动端（iOS/Android）操作，包括审批、查询等核心功能。\nFR-005: 报表中心需提供至少15种预置分析报表，支持自定义报表设计。")
    pdf.section("三、非功能需求")
    pdf.body("NFR-001: 系统可用性 ≥ 99.9%，年停机时间不超过8小时。\nNFR-002: 页面响应时间 ≤ 2秒（95分位）。\nNFR-003: 支持并发用户数 ≥ 500。\nNFR-004: 数据存储需满足等保三级要求，关键数据加密存储。\nNFR-005: 系统需支持水平扩展，可根据业务增长动态增加节点。")
    pdf.section("四、集成需求")
    pdf.body("1. 需与现有 ERP 系统（SAP S/4HANA）对接。\n2. 需与财务系统（用友 NC Cloud）对接。\n3. 需与钉钉/飞书审批流集成。")
    pdf.section("五、实施与培训")
    pdf.body("1. 分三期实施：一期核心采购+库存（3个月），二期物流+报表（2个月），三期移动端+AI优化（2个月）。\n2. 需提供管理员培训（不少于16课时）和最终用户培训（不少于8课时）。")
    path = OUT_DIR / "urs_requirements_test.pdf"
    pdf.output(str(path))
    print(f"  ✅ {path.name}")


def generate_credit_contract():
    """生成担保/授信合同 [116] 用"""
    pdf = TestPDF("担保合同 & 授信合同")
    pdf.add_page()
    pdf.section("担保合同")
    pdf.table_row("担保人", "北京神州数码有限公司")
    pdf.table_row("被担保人", "北京元力科技有限公司")
    pdf.table_row("债权人", "中国工商银行北京分行")
    pdf.table_row("担保金额", "人民币 壹仟万元整 (¥10,000,000.00)")
    pdf.table_row("担保期限", "2026年7月1日 至 2029年6月30日")
    pdf.table_row("担保方式", "连带责任保证")
    pdf.ln(8)
    pdf.section("授信合同")
    pdf.table_row("授信银行", "中国工商银行北京分行")
    pdf.table_row("授信客户", "神州数码（中国）有限公司")
    pdf.table_row("授信额度", "人民币 伍亿元整 (¥500,000,000.00)")
    pdf.table_row("授信期限", "2026年1月1日 至 2027年12月31日")
    pdf.table_row("利率", "LPR+0.5%")
    pdf.table_row("用途", "日常经营周转")
    path = OUT_DIR / "credit_contract_test.pdf"
    pdf.output(str(path))
    print(f"  ✅ {path.name}")


def generate_quote_document():
    """生成报价单 [98] 报价单审核 用"""
    pdf = TestPDF("报价单")
    pdf.add_page()
    pdf.section("报价单信息")
    pdf.table_row("报价单号", "Q-2026-0709-015")
    pdf.table_row("报价日期", "2026年7月9日")
    pdf.table_row("有效期至", "2026年8月8日")
    pdf.table_row("报价单位", "神州数码（中国）有限公司")
    pdf.table_row("客户名称", "北京未来科技有限公司")
    pdf.table_row("联系人", "张三 / 13800138000")
    pdf.ln(3)
    pdf.section("报价明细")
    pdf.body("┌──────┬────────────────────┬────┬──────────┬─────────────┐\n│ 序号 │ 产品名称            │数量│ 单价(元)  │ 小计(元)    │\n├──────┼────────────────────┼────┼──────────┼─────────────┤\n│  1   │ 服务器 DELL R750xs  │ 5  │ 45,000   │ 225,000     │\n│  2   │ 交换机 H3C S5560X   │ 10 │ 8,500    │ 85,000      │\n│  3   │ 防火墙 山石 SG-6000 │ 2  │ 32,000   │ 64,000      │\n│  4   │ 实施部署服务        │ 1  │ 50,000   │ 50,000      │\n├──────┼────────────────────┼────┼──────────┼─────────────┤\n│      │ 合计                │    │          │ 424,000     │\n└──────┴────────────────────┴────┴──────────┴─────────────┘")
    pdf.table_row("总金额(大写)", "人民币 肆拾贰万肆仟元整")
    pdf.ln(3)
    pdf.section("备注")
    pdf.body("1. 以上报价含增值税（税率13%）。\n2. 付款方式：签约后预付30%，到货验收后支付70%。\n3. 交货期：合同签订后30个工作日。\n4. 质保期：硬件3年，实施服务1年。")
    pdf.section("签章")
    pdf.body("报价单位（盖章）：神州数码（中国）有限公司\n日期：2026年7月9日")
    path = OUT_DIR / "quote_document_test.pdf"
    pdf.output(str(path))
    print(f"  ✅ {path.name}")


def generate_sensitive_pdf():
    """含敏感信息的 PDF [61] PDF脱敏打码 用"""
    pdf = TestPDF("员工信息登记表")
    pdf.add_page()
    pdf.section("个人信息")
    pdf.table_row("姓名", "张伟")
    pdf.table_row("身份证号", "110101199001011234")
    pdf.table_row("手机号码", "13800138000")
    pdf.table_row("银行卡号", "6222020200012345678")
    pdf.table_row("电子邮箱", "zhangwei@example.com")
    pdf.table_row("家庭住址", "北京市海淀区中关村南大街5号院1号楼302室")
    pdf.ln(5)
    pdf.section("工作信息")
    pdf.table_row("员工编号", "DC20200001")
    pdf.table_row("部门", "技术研发中心")
    pdf.table_row("职位", "高级软件工程师")
    pdf.table_row("入职日期", "2020年3月15日")
    pdf.ln(5)
    pdf.section("紧急联系人")
    pdf.table_row("姓名", "李芳")
    pdf.table_row("关系", "配偶")
    pdf.table_row("手机号码", "13900139000")
    pdf.body("声明：本人确认以上信息真实有效，授权公司用于人事管理及薪酬发放用途。\n签字：________  日期：________")
    path = OUT_DIR / "employee_info_sensitive.pdf"
    pdf.output(str(path))
    print(f"  ✅ {path.name}")


def generate_enterprise_form():
    """生成企业信息表格 (Excel) [112] 企业信息收集表格自动填写 用"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商信息调查表"

    # 填写样例数据（部分字段）
    data = [
        ["供应商信息调查表", "", ""],
        ["", "", ""],
        ["企业基本信息", "", ""],
        ["企业全称", "北京元力科技有限公司", ""],
        ["统一社会信用代码", "", ""],
        ["注册地址", "", ""],
        ["法定代表人", "", ""],
        ["注册资本", "", ""],
        ["成立日期", "", ""],
        ["企业类型", "有限责任公司", ""],
        ["", "", ""],
        ["联系人信息", "", ""],
        ["联系人姓名", "王芳", ""],
        ["联系电话", "", ""],
        ["电子邮箱", "", ""],
        ["", "", ""],
        ["资质信息", "", ""],
        ["ISO9001认证", "", ""],
        ["ISO27001认证", "", ""],
        ["高新技术企业", "是", ""],
        ["", "", ""],
        ["主要客户", "", ""],
        ["客户名称1", "神州数码（中国）有限公司", ""],
        ["合作年限", "5年", ""],
        ["客户名称2", "", ""],
        ["合作年限", "", ""],
    ]
    for row in data:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20

    path = OUT_DIR / "enterprise_info_form.xlsx"
    wb.save(str(path))
    print(f"  ✅ {path.name}")


if __name__ == "__main__":
    print("生成测试文件...")
    generate_sales_contract()
    generate_purchase_contract()
    generate_urs_document()
    generate_credit_contract()
    generate_quote_document()
    generate_sensitive_pdf()
    generate_enterprise_form()
    print(f"\n📁 全部文件: {OUT_DIR}")
